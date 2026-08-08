from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = PROJECT_ROOT / "data" / "madden_27_launch_ratings.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "madden_27_players.csv"
DEFAULT_METADATA = PROJECT_ROOT / "data" / "madden_27_metadata.json"
DEFAULT_RAW = PROJECT_ROOT / "data" / "madden_27_raw.json"

TEAM_NORMALIZATION = {
    "NY Giants": "New York Giants",
    "NY Jets": "New York Jets",
}

POSITION_NORMALIZATION = {
    "LEDG": "EDGE",
    "REDG": "EDGE",
    "MIKE": "LB",
    "WILL": "LB",
    "SAM": "LB",
}

COLUMN_MAP = {
    "id": "ea_player_id",
    "birthdate": "birthdate",
    "height": "height",
    "weight": "weight",
    "overallRating": "overall",
    "college": "college",
    "handedness": "handedness",
    "age": "age",
    "jerseyNum": "jersey_number",
    "yearsPro": "years_pro",
    "archetype/label": "archetype",
    "team/label": "team",
    "position/shortLabel": "madden_position_raw",
    "position/label": "madden_position_label",
    "position/positionType/name": "madden_position_type",
    "iteration/label": "madden_iteration",
    "stats/acceleration/value": "acceleration",
    "stats/agility/value": "agility",
    "stats/jumping/value": "jumping",
    "stats/stamina/value": "stamina",
    "stats/strength/value": "strength",
    "stats/awareness/value": "awareness",
    "stats/bCVision/value": "bc_vision",
    "stats/blockShedding/value": "block_shedding",
    "stats/breakSack/value": "break_sack",
    "stats/breakTackle/value": "break_tackle",
    "stats/carrying/value": "carrying",
    "stats/catchInTraffic/value": "catch_in_traffic",
    "stats/catching/value": "catching",
    "stats/changeOfDirection/value": "change_of_direction",
    "stats/deepRouteRunning/value": "deep_route_running",
    "stats/finesseMoves/value": "finesse_moves",
    "stats/hitPower/value": "hit_power",
    "stats/impactBlocking/value": "impact_blocking",
    "stats/injury/value": "injury",
    "stats/jukeMove/value": "juke_move",
    "stats/kickAccuracy/value": "kick_accuracy",
    "stats/kickPower/value": "kick_power",
    "stats/kickReturn/value": "kick_return",
    "stats/leadBlock/value": "lead_block",
    "stats/manCoverage/value": "man_coverage",
    "stats/mediumRouteRunning/value": "medium_route_running",
    "stats/passBlock/value": "pass_block",
    "stats/passBlockFinesse/value": "pass_block_finesse",
    "stats/passBlockPower/value": "pass_block_power",
    "stats/playAction/value": "play_action",
    "stats/playRecognition/value": "play_recognition",
    "stats/powerMoves/value": "power_moves",
    "stats/press/value": "press",
    "stats/pursuit/value": "pursuit",
    "stats/release/value": "release",
    "stats/runBlock/value": "run_block",
    "stats/runBlockFinesse/value": "run_block_finesse",
    "stats/runBlockPower/value": "run_block_power",
    "stats/shortRouteRunning/value": "short_route_running",
    "stats/spectacularCatch/value": "spectacular_catch",
    "stats/speed/value": "speed",
    "stats/spinMove/value": "spin_move",
    "stats/stiffArm/value": "stiff_arm",
    "stats/tackle/value": "tackle",
    "stats/throwAccuracyDeep/value": "throw_accuracy_deep",
    "stats/throwAccuracyMid/value": "throw_accuracy_mid",
    "stats/throwAccuracyShort/value": "throw_accuracy_short",
    "stats/throwOnTheRun/value": "throw_on_the_run",
    "stats/throwPower/value": "throw_power",
    "stats/throwUnderPressure/value": "throw_under_pressure",
    "stats/toughness/value": "toughness",
    "stats/trucking/value": "trucking",
    "stats/zoneCoverage/value": "zone_coverage",
}


def _atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + ".tmp")
    temp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    temp.replace(path)


def load_launch_workbook(path: Path | str = DEFAULT_WORKBOOK) -> pd.DataFrame:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if "Launch Ratings" not in workbook.sheetnames:
        raise RuntimeError("Madden workbook is missing the 'Launch Ratings' sheet.")
    ws = workbook["Launch Ratings"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    missing = [column for column in ("firstName", "lastName", "overallRating", "team/label", "position/shortLabel") if column not in headers]
    if missing:
        raise RuntimeError("Madden launch workbook is missing required columns: " + ", ".join(missing))

    records = [row for row in rows if any(value is not None for value in row)]
    raw = pd.DataFrame(records, columns=headers)
    out = pd.DataFrame(index=raw.index)
    out["player_name"] = (raw["firstName"].fillna("").astype(str).str.strip() + " " + raw["lastName"].fillna("").astype(str).str.strip()).str.strip()
    for source, target in COLUMN_MAP.items():
        if source in raw.columns:
            out[target] = raw[source]

    out["team"] = out["team"].replace(TEAM_NORMALIZATION).fillna("").astype(str).str.strip()
    raw_pos = out["madden_position_raw"].fillna("").astype(str).str.upper().str.strip()
    out["position"] = raw_pos.replace(POSITION_NORMALIZATION)
    out["madden_position_raw"] = raw_pos
    out["madden_source"] = "Madden NFL 27 Launch Ratings workbook"

    numeric = [column for column in out.columns if column not in {
        "player_name", "team", "position", "madden_position_raw", "madden_position_label",
        "madden_position_type", "madden_iteration", "archetype", "college", "handedness", "birthdate", "madden_source"
    }]
    for column in numeric:
        out[column] = pd.to_numeric(out[column], errors="coerce")

    out = out[out["player_name"].ne("") & out["team"].ne("") & out["position"].ne("") & out["overall"].notna()].copy()
    out["overall"] = out["overall"].clip(0, 99)
    out = out.drop_duplicates(subset=["ea_player_id"], keep="first")
    return out.reset_index(drop=True)


def import_launch_ratings(
    workbook_path: Path | str = DEFAULT_WORKBOOK,
    output_path: Path | str = DEFAULT_OUTPUT,
    metadata_path: Path | str = DEFAULT_METADATA,
    raw_path: Path | str = DEFAULT_RAW,
) -> pd.DataFrame:
    players = load_launch_workbook(workbook_path)
    teams = sorted(players["team"].dropna().unique().tolist())
    if len(teams) != 32:
        raise RuntimeError(f"Expected 32 NFL teams in Madden launch ratings; found {len(teams)}.")
    if len(players) < 2000:
        raise RuntimeError(f"Expected a full Madden launch roster; found only {len(players)} players.")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp = output.with_suffix(output.suffix + ".tmp")
    players.to_csv(temp, index=False)
    temp.replace(output)

    metadata = {
        "source": "Madden NFL 27 Launch Ratings workbook",
        "source_file": str(Path(workbook_path)),
        "sheet": "Launch Ratings",
        "imported_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "players": int(len(players)),
        "teams": len(teams),
        "columns": int(len(players.columns)),
        "iteration_labels": sorted({str(x) for x in players.get("madden_iteration", pd.Series(dtype=object)).dropna().unique()}),
        "authoritative_preseason_source": True,
        "ea_live_download_disabled_for_this_refresh": True,
    }
    _atomic_json(Path(metadata_path), metadata)
    _atomic_json(Path(raw_path), {
        "source": "Madden NFL 27 Launch Ratings workbook",
        "source_file": str(Path(workbook_path)),
        "sheet": "Launch Ratings",
        "players": int(len(players)),
        "teams": len(teams),
        "note": "The authoritative raw source is stored in data/madden_27_launch_ratings.xlsx; this JSON is provenance only.",
    })
    return players


if __name__ == "__main__":
    frame = import_launch_ratings()
    print(f"Imported {len(frame)} Madden 27 launch-rating players across {frame['team'].nunique()} teams.")
